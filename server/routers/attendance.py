"""Attendance endpoints: check-in / check-out sessions, mark attendance, and export spreadsheet."""

from __future__ import annotations

import io
from datetime import date, datetime

from fastapi import APIRouter, Depends, HTTPException, Query, Response
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from dependencies import get_current_staff
from models import (
    AttendanceByIdRecord,
    AttendanceRecord,
    CheckInRequest,
    CheckOutRequest,
    SessionResponse,
    SessionsForDateResponse,
)
from services.attendance_service import (
    AlreadyCheckedInError,
    AlreadyCheckedOutError,
    AttendanceService,
    DepartmentNotFoundError,
    RegistrantNotFoundError,
    service as attendance_service,
)
from services.department_service import (
    DepartmentService,
    service as department_service,
)
from services.staff_auth import StaffPublic

router = APIRouter(tags=["attendance"])


@router.post("/departments/{department_code}/check-in", status_code=201)
async def check_in(
    department_code: str,
    body: CheckInRequest,
    staff: StaffPublic = Depends(get_current_staff),
):
    try:
        occurred_at = datetime.fromisoformat(body.occurred_at) if body.occurred_at else None
        result = attendance_service.check_in(
            department_code=department_code,
            registrant_id=body.id,
            occurred_at=occurred_at,
            staff_id=None,
        )
        return result
    except DepartmentNotFoundError as exc:
        raise HTTPException(status_code=404, detail=str(exc))
    except RegistrantNotFoundError as exc:
        raise HTTPException(status_code=404, detail=str(exc))
    except AlreadyCheckedInError as exc:
        raise HTTPException(status_code=400, detail=str(exc))


@router.post("/departments/{department_code}/check-out", status_code=201)
async def check_out(
    department_code: str,
    body: CheckOutRequest,
    staff: StaffPublic = Depends(get_current_staff),
):
    try:
        occurred_at = datetime.fromisoformat(body.occurred_at) if body.occurred_at else None
        result = attendance_service.check_out(
            department_code=department_code,
            registrant_id=body.id,
            occurred_at=occurred_at,
            staff_id=None,
        )
        return result
    except DepartmentNotFoundError as exc:
        raise HTTPException(status_code=404, detail=str(exc))
    except RegistrantNotFoundError as exc:
        raise HTTPException(status_code=404, detail=str(exc))
    except AlreadyCheckedOutError as exc:
        raise HTTPException(status_code=400, detail=str(exc))


@router.get("/departments/{department_code}/attendance/sessions")
async def get_sessions_for_date(
    department_code: str,
    date: date = Query(default=..., description="Date in YYYY-MM-DD format"),
    staff: StaffPublic = Depends(get_current_staff),
):
    try:
        sessions = attendance_service.get_sessions_for_date(department_code, date)
        return {"date": date, "sessions": sessions}
    except DepartmentNotFoundError as exc:
        raise HTTPException(status_code=404, detail=str(exc))


@router.post("/departments/{department_code}/attendance/{matric_number}")
async def mark_attendance(
    department_code: str,
    matric_number: str,
    record: AttendanceRecord,
    staff: StaffPublic = Depends(get_current_staff),
):
    raise HTTPException(
        status_code=410,
        detail="This endpoint is deprecated. Use POST /departments/{code}/check-in instead.",
    )


# Deprecated: use POST /departments/{code}/check-in instead.
@router.post("/departments/{department_code}/attendance", status_code=200)
async def mark_attendance_by_id(
    department_code: str,
    record: AttendanceByIdRecord,
    staff: StaffPublic = Depends(get_current_staff),
):
    raise HTTPException(
        status_code=410,
        detail="This endpoint is deprecated. Use POST /departments/{code}/check-in instead.",
    )


@router.get("/departments/{department_code}/attendance/spreadsheet")
async def get_attendance_spreadsheet(
    department_code: str,
    staff: StaffPublic = Depends(get_current_staff),
):
    try:
        data = attendance_service.get_attendance_spreadsheet(department_code)
        department = department_service._get_department(department_code)

        REGISTRANT_KEYS = {"id", "name", "email", "phone", "matriculation_number"}
        date_columns: list[str] = sorted(
            {k for row in data for k in row if k not in REGISTRANT_KEYS}
        )

        wb = Workbook()
        ws = wb.active
        ws.title = "Attendance"

        HEADER_FILL   = PatternFill("solid", fgColor="4A0072")
        DATE_FILL     = PatternFill("solid", fgColor="6A1B9A")
        META_FILL     = PatternFill("solid", fgColor="F3E5F5")
        WHITE_FONT    = Font(color="FFFFFF", bold=True, name="Calibri", size=11)
        BOLD_FONT     = Font(bold=True, name="Calibri", size=11)
        NORMAL_FONT   = Font(name="Calibri", size=10)
        CENTER        = Alignment(horizontal="center", vertical="center", wrap_text=True)
        LEFT          = Alignment(horizontal="left",   vertical="center", wrap_text=True)
        THIN          = Side(style="thin", color="CCCCCC")
        THIN_BORDER   = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

        def _style(cell, font=None, fill=None, align=None, border=None):
            if font:   cell.font   = font
            if fill:   cell.fill   = fill
            if align:  cell.alignment = align
            if border: cell.border = border

        total_cols = 5 + len(date_columns)
        last_col   = get_column_letter(max(total_cols, 1))

        ws.merge_cells(f"A1:{last_col}1")
        title_cell = ws["A1"]
        title_cell.value = f"{department.get('name', department_code)} — Attendance Register"
        _style(title_cell,
               font=Font(bold=True, size=14, color="4A0072", name="Calibri"),
               fill=META_FILL, align=LEFT)
        ws.row_dimensions[1].height = 28

        meta_pairs = [
            ("Department Code",  department.get("code", department_code)),
            ("Duration",     department.get("duration", "")),
            ("Generated",    datetime.now().strftime("%Y-%m-%d %H:%M")),
        ]
        for r_offset, (label, value) in enumerate(meta_pairs, start=2):
            lc = ws.cell(row=r_offset, column=1, value=label)
            vc = ws.cell(row=r_offset, column=2, value=value)
            _style(lc, font=BOLD_FONT,   fill=META_FILL, align=LEFT)
            _style(vc, font=NORMAL_FONT, fill=META_FILL, align=LEFT)
            if total_cols > 2:
                ws.merge_cells(
                    start_row=r_offset, start_column=2,
                    end_row=r_offset,   end_column=total_cols
                )

        ws.row_dimensions[5].height = 8

        HEADER_ROW = 6
        base_defs = [
            ("#",              None,                     5),
            ("Name",           "name",                  28),
            ("Email",          "email",                 34),
            ("Phone",          "phone",                 16),
            ("Matric No",      "matriculation_number",  18),
        ]

        for col_idx, (label, _, width) in enumerate(base_defs, start=1):
            cell = ws.cell(row=HEADER_ROW, column=col_idx, value=label)
            _style(cell, font=WHITE_FONT, fill=HEADER_FILL,
                   align=CENTER, border=THIN_BORDER)
            ws.column_dimensions[get_column_letter(col_idx)].width = width

        for d_idx, date_str in enumerate(date_columns):
            col_idx = len(base_defs) + 1 + d_idx
            cell = ws.cell(row=HEADER_ROW, column=col_idx, value=date_str)
            _style(cell, font=WHITE_FONT, fill=DATE_FILL,
                   align=CENTER, border=THIN_BORDER)
            ws.column_dimensions[get_column_letter(col_idx)].width = 12

        ws.row_dimensions[HEADER_ROW].height = 36

        PRESENT_FILL  = PatternFill("solid", fgColor="E8F5E9")
        ABSENT_FILL   = PatternFill("solid", fgColor="FFEBEE")
        PARTIAL_FILL  = PatternFill("solid", fgColor="FFF8E1")
        ALT_FILL      = PatternFill("solid", fgColor="FAF5FF")

        for row_offset, row in enumerate(data):
            excel_row = HEADER_ROW + 1 + row_offset
            row_fill  = ALT_FILL if row_offset % 2 == 1 else None

            cell = ws.cell(row=excel_row, column=1, value=row_offset + 1)
            _style(cell, font=NORMAL_FONT, fill=row_fill, align=CENTER, border=THIN_BORDER)

            for col_idx, (_, db_key, _) in enumerate(base_defs[1:], start=2):
                value = row.get(db_key, "")
                cell  = ws.cell(row=excel_row, column=col_idx, value=value)
                _style(cell, font=NORMAL_FONT, fill=row_fill, align=LEFT, border=THIN_BORDER)

            for d_idx, date_str in enumerate(date_columns):
                col_idx = len(base_defs) + 1 + d_idx
                raw     = row.get(date_str)
                if raw is None:
                    display, cell_fill = "–", row_fill
                elif raw == 1:
                    display, cell_fill = 1, PRESENT_FILL
                elif raw == 0:
                    display, cell_fill = 0, ABSENT_FILL
                else:
                    display, cell_fill = "P", PARTIAL_FILL

                cell = ws.cell(row=excel_row, column=col_idx, value=display)
                _style(cell, font=NORMAL_FONT, fill=cell_fill, align=CENTER, border=THIN_BORDER)

        ws.freeze_panes = ws.cell(row=HEADER_ROW + 1, column=3)

        if not data:
            note_row = HEADER_ROW + 1
            ws.merge_cells(
                start_row=note_row, start_column=1,
                end_row=note_row,   end_column=max(total_cols, 1)
            )
            note = ws.cell(row=note_row, column=1,
                               value="No registrants found for this department.")
            _style(note, font=Font(italic=True, color="888888", name="Calibri"),
                   align=CENTER)

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)

        safe_code = department_code.replace("/", "_")
        return Response(
            content=buf.getvalue(),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={
                "Content-Disposition": f'attachment; filename="{safe_code}_attendance.xlsx"',
            },
        )
    except DepartmentNotFoundError as exc:
        raise HTTPException(status_code=404, detail=str(exc))
    except RegistrantNotFoundError as exc:
        raise HTTPException(status_code=404, detail=str(exc))
