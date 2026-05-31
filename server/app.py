import logging
import os
import io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import date
from datetime import datetime, timedelta, timezone
from pathlib import Path
from uuid import uuid4
from dotenv import load_dotenv

# Load .env from the server directory during local development so
# environment variables defined in .env are available via os.getenv().
load_dotenv(dotenv_path=Path(__file__).parent / ".env")

from fastapi import FastAPI, HTTPException, Depends, File, Form, UploadFile
from fastapi.middleware.cors import CORSMiddleware
from fastapi.staticfiles import StaticFiles
from fastapi.responses import StreamingResponse, Response
from pydantic import BaseModel, EmailStr
from fastapi.security import OAuth2PasswordRequestForm

from course_service import (
    AttendanceAlreadyMarkedError,
    CourseAlreadyExistsError,
    CourseNotFoundError,
    CourseService,
    RegistrantExistsError,
    RegistrantNotFoundError,
)

from email_service import email_service
from staff_auth import (
    StaffAuthService,
    StaffAlreadyExistsError,
    StaffNotVerifiedError,
    StaffPublic,
    StaffNotFoundError,
    InvalidCredentialsError,
    InvalidVerificationCodeError,
    VerificationCodeExpiredError,
    oauth2_scheme,
)

app = FastAPI()

# Create uploads directory if it doesn't exist
from db import UPLOAD_DIR
UPLOAD_DIR.mkdir(exist_ok=True)


# Add CORS middleware to allow requests from mobile app
app.add_middleware(
    CORSMiddleware,
    allow_origins=[
        "http://localhost:8081",  # Web
        "http://localhost:8100",  # Web (alt)
        "http://127.0.0.1:8081",
        "http://10.0.2.2:8081",   # Android emulator (points to host)
        "http://10.1.1.240:8081", # Expo LAN host IP + dev server port
        "http://10.1.1.240:8000", # Expo LAN host IP + API port
        "*",  # For development; restrict in production
    ],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# Serve uploaded files as static files
app.mount("/uploads", StaticFiles(directory="uploads"), name="uploads")


service = CourseService()
auth_service = StaffAuthService(secret_key=os.getenv("JWT_SECRET", "dev-secret-change-me"))

class AttendanceRecord(BaseModel):
    date: date
    present: bool = True


class RegistrantCreate(BaseModel):
    name: str
    email: EmailStr
    phone: str
    matriculation_number: str


class LoginResponse(BaseModel):
    access_token: str
    token_type: str = "bearer"


class StaffRegisterRequest(BaseModel):
    email: EmailStr
    username: str
    password: str


class StaffVerifyRequest(BaseModel):
    username: str
    pin: str


class StaffRegisterResponse(BaseModel):
    message: str
    username: str
    email: EmailStr
    verification_expires_at: datetime


def get_current_staff(token: str = Depends(oauth2_scheme)) -> StaffPublic:
    try:
        payload = auth_service.decode_access_token(token)
        username = payload.get("sub")
        if not username:
            raise InvalidCredentialsError("Invalid token subject")

        staff = auth_service.find_staff_by_username(username)
        if not staff:
            raise StaffNotFoundError("Staff account not found")

        return StaffPublic(
            username=staff["username"],
            name=staff["name"],
            email=staff["email"],
        )
    except (InvalidCredentialsError, StaffNotFoundError):
        raise HTTPException(status_code=401, detail="Could not validate credentials")


@app.get("/")
async def root():
    return {"message": "Attendance API"}


@app.get("/health")
async def health():
    return {"status": "ok"}


class CourseCreate(BaseModel):
    code: str
    name: str
    description: str = ""
    duration: str = ""


@app.post("/courses", status_code=201)
async def create_course(
    payload: CourseCreate,
    staff: StaffPublic = Depends(get_current_staff),
):
    try:
        course = service.create_course(
            code=payload.code.strip().upper(),
            name=payload.name.strip(),
            description=payload.description.strip(),
            duration=payload.duration.strip(),
        )
        return {"message": "Event created", "course": course}
    except CourseAlreadyExistsError as e:
        raise HTTPException(status_code=400, detail=str(e))


@app.get("/courses")
async def get_courses():
    courses = service.list_courses()
    return {
        "courses": [
            {
                "name": course.get("name") or course.get("course_name") or "",
                "code": course.get("code") or course.get("course_code") or "",
                "description": course.get("description", ""),
                "duration": course.get("duration", ""),
            }
            for course in courses
        ]
    }


@app.get("/courses/{course_code}/registrants")
async def get_course_registrants(
    course_code: str,
    staff: StaffPublic = Depends(get_current_staff),
):
    try:
        return {
            "code": course_code,
            "registrants": service.get_registrants(course_code),
        }
    except CourseNotFoundError as e:
        raise HTTPException(status_code=404, detail=str(e))


@app.get("/courses/{course_code}/registrants/{registrant_id}")
async def get_course_registrant(
    course_code: str,
    registrant_id: str,
    staff: StaffPublic = Depends(get_current_staff),
):
    try:
        return {
            "code": course_code,
            "registrant": service.get_registrant(course_code, registrant_id),
        }
    except CourseNotFoundError as e:
        raise HTTPException(status_code=404, detail=str(e))
    except RegistrantNotFoundError as e:
        raise HTTPException(status_code=404, detail=str(e))


@app.get("/courses/{course_code}/scan-context")
async def get_course_scan_context(course_code: str):
    try:
        return service.get_scan_context(course_code)
    except CourseNotFoundError as e:
        raise HTTPException(status_code=404, detail=str(e))
    except RegistrantNotFoundError as e:
        raise HTTPException(status_code=404, detail=str(e))


@app.post("/courses/{course_code}/register")
async def register_for_course(
    course_code: str,
    name: str = Form(...),
    email: str = Form(...),
    phone: str = Form(...),
    matriculation_number: str = Form(...),
    image: UploadFile | None = File(None),
):
    try:
        image_filename = None
        
        # Save uploaded image if provided
        if image:
            # Generate unique filename
            file_extension = Path(image.filename).suffix if image.filename else ".jpg"
            image_filename = f"{uuid4().hex}{file_extension}"
            
            # Save file to uploads directory
            file_path = UPLOAD_DIR / image_filename
            content = await image.read()
            with open(file_path, "wb") as f:
                f.write(content)
        
        # Register the attendee
        registrant_data = {
            "name": name,
            "email": email,
            "phone": phone,
            "matriculation_number": matriculation_number,
            "image_url": f"/uploads/{image_filename}" if image_filename else None,
        }
        
        created = service.register(course_code, registrant_data)

        # Fetch course info to include in registration email
        try:
            course_info = service._get_course(course_code)
        except CourseNotFoundError:
            course_info = None

        # Send registration email with QR code - rollback on failure
        try:
            email_service.send_registration_email(created, created.get("qr_bytes", b""), course=course_info)
        except Exception as exc:
            logging.getLogger(__name__).warning(
                "Failed to send registration email to %s: %s",
                created.get("email"),
                exc,
            )
            # Rollback: delete the registrant and image
            service.delete_registrant(course_code, created["id"])
            if image_filename:
                file_path = UPLOAD_DIR / image_filename
                if file_path.exists():
                    file_path.unlink()
            raise HTTPException(status_code=500, detail="Registration failed: could not send confirmation email")

        # Return only a success message — no personal data echoed back
        return {"message": "Registration saved"}
    except CourseNotFoundError as e:
        raise HTTPException(status_code=404, detail=str(e))
    except RegistrantExistsError as e:
        raise HTTPException(status_code=400, detail=str(e))


@app.post("/courses/{course_code}/attendance/{matric_number}")
async def mark_attendance(
    course_code: str,
    matric_number: str,
    record: AttendanceRecord,
    staff: StaffPublic = Depends(get_current_staff),
):
    try:
        attendance = service.mark_attendance(
            course_code=course_code,
            matric_number=matric_number,
            attendance_date=record.date,
            present=record.present,
        )
        return {"message": "Attendance recorded", "attendance": attendance}
    except CourseNotFoundError as e:
        raise HTTPException(status_code=404, detail=str(e))
    except RegistrantNotFoundError as e:
        raise HTTPException(status_code=404, detail=str(e))
    except AttendanceAlreadyMarkedError as e:
        raise HTTPException(status_code=400, detail=str(e))


@app.get("/courses/{course_code}/attendance/spreadsheet")
async def get_attendance_spreadsheet(
    course_code: str,
    staff: StaffPublic = Depends(get_current_staff),
):
    try:
        # ── 1. Raw data ─────────────────────────────────────────────────────────
        data = service.get_attendance_spreadsheet(course_code)
        course = service._get_course(course_code)

        # Collect all unique date columns (anything that's not a registrant field)
        REGISTRANT_KEYS = {"id", "name", "email", "phone", "matriculation_number"}
        date_columns: list[str] = sorted(
            {k for row in data for k in row if k not in REGISTRANT_KEYS}
        )

        # ── 2. Build workbook ────────────────────────────────────────────────────
        wb = Workbook()
        ws = wb.active
        ws.title = "Attendance"

        # ── Shared style helpers ─────────────────────────────────────────────────
        HEADER_FILL   = PatternFill("solid", fgColor="4A0072")   # deep purple
        DATE_FILL     = PatternFill("solid", fgColor="6A1B9A")   # medium purple
        META_FILL     = PatternFill("solid", fgColor="F3E5F5")   # light lavender
        WHITE_FONT    = Font(color="FFFFFF", bold=True, name="Calibri", size=11)
        BOLD_FONT     = Font(bold=True, name="Calibri", size=11)
        NORMAL_FONT   = Font(name="Calibri", size=10)
        CENTER        = Alignment(horizontal="center", vertical="center", wrap_text=True)
        LEFT          = Alignment(horizontal="left",   vertical="center", wrap_text=True)
        THIN          = Side(style="thin", color="CCCCCC")
        THIN_BORDER   = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

        def _style(cell, font=None, fill=None, align=None, border=None):
            if font:   cell.font   = font
            if fill:   cell.fill   = fill
            if align:  cell.alignment = align
            if border: cell.border = border

        # ── 3. Metadata block (rows 1-4) ─────────────────────────────────────────
        total_cols = 5 + len(date_columns)                       # A … last-date col
        last_col   = get_column_letter(max(total_cols, 1))

        ws.merge_cells(f"A1:{last_col}1")
        title_cell = ws["A1"]
        title_cell.value = f"{course.get('name', course_code)} — Attendance Register"
        _style(title_cell,
               font=Font(bold=True, size=14, color="4A0072", name="Calibri"),
               fill=META_FILL, align=LEFT)
        ws.row_dimensions[1].height = 28

        meta_pairs = [
            ("Course Code",  course.get("code", course_code)),
            ("Description",  course.get("description", "")),
            ("Duration",     course.get("duration", "")),
            ("Generated",    datetime.now().strftime("%Y-%m-%d %H:%M")),
        ]
        for r_offset, (label, value) in enumerate(meta_pairs, start=2):
            lc = ws.cell(row=r_offset, column=1, value=label)
            vc = ws.cell(row=r_offset, column=2, value=value)
            _style(lc, font=BOLD_FONT,   fill=META_FILL, align=LEFT)
            _style(vc, font=NORMAL_FONT, fill=META_FILL, align=LEFT)
            # Merge remaining columns so the row looks clean
            if total_cols > 2:
                ws.merge_cells(
                    start_row=r_offset, start_column=2,
                    end_row=r_offset,   end_column=total_cols
                )

        # Blank separator row
        ws.row_dimensions[6].height = 8

        # ── 4. Column headers (row 7) ─────────────────────────────────────────────
        HEADER_ROW = 7
        base_defs = [
            # (display label, db key, col width)
            ("#",              None,                     5),
            ("Name",           "name",                  28),
            ("Email",          "email",                 34),
            ("Phone",          "phone",                 16),
            ("Matric No",      "matriculation_number",  18),
        ]

        for col_idx, (label, _, width) in enumerate(base_defs, start=1):
            cell = ws.cell(row=HEADER_ROW, column=col_idx, value=label)
            _style(cell, font=WHITE_FONT, fill=HEADER_FILL,
                   align=CENTER, border=THIN_BORDER)
            ws.column_dimensions[get_column_letter(col_idx)].width = width

        for d_idx, date_str in enumerate(date_columns):
            col_idx = len(base_defs) + 1 + d_idx
            cell = ws.cell(row=HEADER_ROW, column=col_idx, value=date_str)
            _style(cell, font=WHITE_FONT, fill=DATE_FILL,
                   align=CENTER, border=THIN_BORDER)
            ws.column_dimensions[get_column_letter(col_idx)].width = 12

        ws.row_dimensions[HEADER_ROW].height = 36

        # ── 5. Data rows ──────────────────────────────────────────────────────────
        PRESENT_FILL  = PatternFill("solid", fgColor="E8F5E9")   # pale green
        ABSENT_FILL   = PatternFill("solid", fgColor="FFEBEE")   # pale red
        ALT_FILL      = PatternFill("solid", fgColor="FAF5FF")   # very light purple

        for row_offset, row in enumerate(data):
            excel_row = HEADER_ROW + 1 + row_offset
            row_fill  = ALT_FILL if row_offset % 2 == 1 else None   # zebra striping

            # Row number
            cell = ws.cell(row=excel_row, column=1, value=row_offset + 1)
            _style(cell, font=NORMAL_FONT, fill=row_fill, align=CENTER, border=THIN_BORDER)

            # Registrant fields
            for col_idx, (_, db_key, _) in enumerate(base_defs[1:], start=2):
                value = row.get(db_key, "")
                cell  = ws.cell(row=excel_row, column=col_idx, value=value)
                _style(cell, font=NORMAL_FONT, fill=row_fill, align=LEFT, border=THIN_BORDER)

            # Attendance day columns
            for d_idx, date_str in enumerate(date_columns):
                col_idx = len(base_defs) + 1 + d_idx
                raw     = row.get(date_str)           # True / False / None (no record)
                if raw is None:
                    display = "–"                     # no record at all
                    cell_fill = row_fill
                elif raw:
                    display, cell_fill = 1, PRESENT_FILL
                else:
                    display, cell_fill = 0, ABSENT_FILL

                cell = ws.cell(row=excel_row, column=col_idx, value=display)
                _style(cell, font=NORMAL_FONT, fill=cell_fill, align=CENTER, border=THIN_BORDER)

        # Freeze panes so header + ID columns stay visible while scrolling
        ws.freeze_panes = ws.cell(row=HEADER_ROW + 1, column=3)

        # ── 6. Empty-data note ────────────────────────────────────────────────────
        if not data:
            note_row = HEADER_ROW + 1
            ws.merge_cells(
                start_row=note_row, start_column=1,
                end_row=note_row,   end_column=max(total_cols, 1)
            )
            note = ws.cell(row=note_row, column=1,
                           value="No registrants found for this course.")
            _style(note, font=Font(italic=True, color="888888", name="Calibri"),
                   align=CENTER)

        # ── 7. Stream response ────────────────────────────────────────────────────
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)

        safe_code = course_code.replace("/", "_")
        return Response(
            content=buf.getvalue(),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={
                "Content-Disposition": f'attachment; filename="{safe_code}_attendance.xlsx"',
            },
        )
    except CourseNotFoundError as e:
        raise HTTPException(status_code=404, detail=str(e))
    except RegistrantNotFoundError as e:
        raise HTTPException(status_code=404, detail=str(e))



@app.delete("/courses/{course_code}", status_code=200)
async def delete_course(
    course_code: str,
    staff: StaffPublic = Depends(get_current_staff),
):
    try:
        service.delete_course(course_code)
        return {"message": "Event deleted"}
    except CourseNotFoundError as e:
        raise HTTPException(status_code=404, detail=str(e))


@app.post("/auth/login", response_model=LoginResponse)
async def login(form_data: OAuth2PasswordRequestForm = Depends()):
    try:
        staff = auth_service.authenticate(form_data.username, form_data.password)
        token = auth_service.create_access_token(subject=staff["username"])
        return {"access_token": token, "token_type": "bearer"}
    except (StaffNotFoundError, InvalidCredentialsError):
        raise HTTPException(status_code=401, detail="Invalid username or password")
    except StaffNotVerifiedError:
        raise HTTPException(status_code=403, detail="Account not verified yet")


@app.post("/auth/register", response_model=StaffRegisterResponse, status_code=201)
async def register_staff(payload: StaffRegisterRequest):
    verification_pin = auth_service.generate_verification_pin()
    verification_expires_at = datetime.now(timezone.utc) + timedelta(minutes=15)

    try:
        staff = auth_service.create_staff(
            username=payload.username,
            email=payload.email,
            password=payload.password,
            verification_pin=verification_pin,
            verification_pin_expires_at=verification_expires_at,
        )

        try:
            email_service.send_staff_verification_email(
                username=staff["username"],
                email=staff["email"],
                verification_pin=verification_pin,
                expires_at=verification_expires_at.isoformat(),
            )
        except Exception as exc:
            logging.getLogger(__name__).warning(
                "Failed to send staff verification email to admin for %s: %s",
                staff.get("email"),
                exc,
            )

        return {
            "message": "Staff account created. Share the PIN with the admin for verification.",
            "username": staff["username"],
            "email": staff["email"],
            "verification_expires_at": verification_expires_at,
        }
    except StaffAlreadyExistsError as e:
        raise HTTPException(status_code=400, detail=str(e))


@app.post("/auth/verify-account", response_model=LoginResponse)
async def verify_staff_account(payload: StaffVerifyRequest):
    try:
        staff = auth_service.verify_staff_account(payload.username, payload.pin)
        token = auth_service.create_access_token(subject=staff["username"])
        return {"access_token": token, "token_type": "bearer"}
    except StaffNotFoundError as e:
        raise HTTPException(status_code=404, detail=str(e))
    except VerificationCodeExpiredError as e:
        raise HTTPException(status_code=400, detail=str(e))
    except InvalidVerificationCodeError as e:
        raise HTTPException(status_code=400, detail=str(e))


